from __future__ import annotations

import json
from pathlib import Path

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None


def grade(transcript, workspace_path):
    root = Path(workspace_path)
    workbook_path = root / "nasa_budget_recovered.xlsx"
    report_path = root / "recovery_report.json"
    records = []

    workbook_exists = workbook_path.exists()
    records.append({
        "id": "workbook_exists",
        "score": 1.0 if workbook_exists else 0.0,
        "weight": 0.15,
        "description": "The repaired workbook file exists.",
        "details": None if workbook_exists else "nasa_budget_recovered.xlsx not found.",
    })

    workbook_valid = False
    remaining_placeholders = None
    if workbook_exists and load_workbook is not None:
        try:
            wb = load_workbook(workbook_path)
            workbook_valid = True
            remaining = 0
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        if isinstance(cell.value, str) and "???" in cell.value:
                            remaining += 1
            remaining_placeholders = remaining
        except Exception as e:
            records.append({
                "id": "workbook_valid",
                "score": 0.0,
                "weight": 0.20,
                "description": "The output is a valid Excel workbook.",
                "details": f"Workbook parse failed: {e}",
            })
    else:
        records.append({
            "id": "workbook_valid",
            "score": 0.0 if not workbook_exists else 0.5,
            "weight": 0.20,
            "description": "The output is a valid Excel workbook.",
            "details": "openpyxl unavailable or workbook missing.",
        })

    if workbook_valid:
        records.append({
            "id": "workbook_valid",
            "score": 1.0,
            "weight": 0.20,
            "description": "The output is a valid Excel workbook.",
            "details": None,
        })

    placeholder_score = 0.0
    placeholder_details = "Workbook not parsed."
    if remaining_placeholders is not None:
        placeholder_score = 1.0 if remaining_placeholders == 0 else 0.0
        placeholder_details = None if remaining_placeholders == 0 else f"{remaining_placeholders} placeholder cell(s) still contain ???"
    records.append({
        "id": "all_placeholders_filled",
        "score": placeholder_score,
        "weight": 0.25,
        "description": "All placeholder cells marked with ??? have been replaced.",
        "details": placeholder_details,
    })

    report_ok = False
    report_details = "recovery_report.json not found."
    if report_path.exists():
        try:
            report = json.loads(report_path.read_text())
            report_ok = (
                isinstance(report, dict)
                and isinstance(report.get("report"), dict)
                and report["report"].get("output_file") == "nasa_budget_recovered.xlsx"
                and report["report"].get("all_missing_cells_filled") is True
                and report["report"].get("workbook_valid") is True
            )
            report_details = None if report_ok else "Report exists but does not match required structure/values."
        except Exception as e:
            report_details = f"recovery_report.json parse failed: {e}"
    records.append({
        "id": "recovery_report",
        "score": 1.0 if report_ok else 0.0,
        "weight": 0.20,
        "description": "The recovery report has the exact required shape and values.",
        "details": report_details,
    })

    workflow_signal = 0.0
    if workbook_exists:
        workflow_signal += 0.5
    if report_ok:
        workflow_signal += 0.5
    records.append({
        "id": "repair_workflow",
        "score": workflow_signal,
        "weight": 0.20,
        "description": "The agent followed a workbook repair workflow and produced both the repaired workbook and report artifact.",
        "details": None if workflow_signal == 1.0 else "Repair workflow appears incomplete or missing one of the required artifacts.",
    })

    return records
